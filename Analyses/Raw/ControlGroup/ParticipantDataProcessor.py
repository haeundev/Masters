import pandas as pd
from openpyxl import load_workbook
import os


class ParticipantDataProcessor:
    def __init__(self, participant_number):
        self.participant_number = participant_number
        self.excel_file = f"{os.getcwd()}/{participant_number}.xlsx"

    def get_sheet(self, sheet_name, sheet_type):
        #print(f"Processing {sheet_name} sheet...")
        df = pd.read_excel(self.excel_file, sheet_name=sheet_name, header=None)
        if sheet_type == "mini":
            df.columns = ["WORD", "IS_CORRECT"]
        elif sheet_type == "training":
            df.columns = ["WORD", "IS_CORRECT", "RESPONSE_TIME"]
        elif sheet_type == "evaluation":
            df.columns = ["WORD", "IS_CORRECT", "NOISE_TYPE", "SPEAKER"]
        #print(f"{sheet_name} sheet processed successfully.")
        return df

    def get_accuracy_average(self, df):
        if df is None:
            return None
        else:
            percentage = df["IS_CORRECT"].value_counts().get(True, 0) / len(df) * 100
            return round(percentage, 2)

    def get_accuracy_of_noise_type(self, df, noise_type):
        if df is None:
            return None
        else:
            noise_df = df[df["NOISE_TYPE"] == noise_type]
            return self.get_accuracy_average(noise_df)

    def get_valid_sessions(self, dfs):
        return [df for df in dfs if df is not None]

    def calculate_accuracy(self, df):
        accuracy_dict = {}
        grouped = df.groupby('WORD')
        for w, group in grouped:
            accuracy = group['IS_CORRECT'].mean()
            accuracy_dict[w] = accuracy
        acc_df = pd.DataFrame(list(accuracy_dict.items()), columns=['WORD', 'ACCURACY'])
        return acc_df

    def track_accuracies_by_words(self, dfs, is_eval=False):
        acc_df = pd.DataFrame()
        session_names = ['PRE', 'MID', 'POST'] if is_eval else [f"Session {i + 1}" for i in range(len(dfs))]

        for df, session_name in zip(dfs, session_names):
            session_accuracy = self.calculate_accuracy(df)
            session_accuracy.columns = ['WORD', session_name]
            if acc_df.empty:
                acc_df = session_accuracy
            else:
                acc_df = pd.merge(acc_df, session_accuracy, on='WORD', how='outer')

        if not acc_df.empty:
            acc_df = acc_df.sort_values(by='WORD', key=lambda x: x.str.lower())

        return acc_df

    def save_accuracy_to_excel(self, df, sheet_name):
        try:
            with pd.ExcelWriter(self.excel_file, mode='a', engine='openpyxl', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
            #print(f"{sheet_name} sheet is saved to {self.excel_file}")
        except ValueError as e:
            if 'Sheet' in str(e) and 'already exists' in str(e):
                with pd.ExcelWriter(self.excel_file, mode='a', engine='openpyxl') as writer:
                    workbook = load_workbook(self.excel_file)
                    if 'Word Accuracies' in workbook.sheetnames:
                        del workbook[sheet_name]
                        workbook.save(self.excel_file)
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        except Exception as e:
            print(f"An error occurred: {e}")

    def get_contrast(self, word):
        contrast_dict = {
            **dict.fromkeys(
                ["rich", "reach", "itch", "each", "sin", "scene", "list", "least", "chip", "cheap", "filled", "field",
                 "grin", "green"], "ɪ vs. iː"),
            **dict.fromkeys(
                ["bet", "bat", "pet", "pat", "met", "mat", "set", "sat", "ten", "tan", "men", "man", "Ken", "can"],
                "ɛ vs. æ"),
            **dict.fromkeys(
                ["cut", "cot", "but", "bot", "hut", "hot", "nut", "not", "sub", "sob", "fund", "fond", "pup", "pop"],
                "ʌ vs. ɑ"),
            **dict.fromkeys(
                ["look", "Luke", "pull", "pool", "full", "fool", "should", "shooed", "bull", "Boole", "could", "cooed",
                 "would", "wooed"], "ʊ vs. uː")
        }

        return contrast_dict.get(word, "Word not found in the table")

    def calculate_contrast_accuracy(self, df):
        contrast_dict = {}
        grouped = df.groupby('WORD')
        for w, group in grouped:
            contrast = self.get_contrast(w)
            accuracy = group['IS_CORRECT'].mean()
            if contrast in contrast_dict:
                contrast_dict[contrast].append(accuracy)
            else:
                contrast_dict[contrast] = [accuracy]

        contrast_accuracy_dict = {}
        for contrast, accuracies in contrast_dict.items():
            contrast_accuracy_dict[contrast] = sum(accuracies) / len(accuracies)

        acc_df = pd.DataFrame(list(contrast_accuracy_dict.items()), columns=['CONTRAST', 'ACCURACY'])
        return acc_df

    def track_contrast_accuracies_by_words(self, dfs):
        acc_df = pd.DataFrame()
        session_names = [f"Session {i + 1}" for i in range(len(dfs))]

        for df, session_name in zip(dfs, session_names):
            session_accuracy = self.calculate_contrast_accuracy(df)
            session_accuracy.columns = ['CONTRAST', session_name]
            if acc_df.empty:
                acc_df = session_accuracy
            else:
                acc_df = pd.merge(acc_df, session_accuracy, on='CONTRAST', how='outer')

        if not acc_df.empty:
            acc_df = acc_df.sort_values(by='CONTRAST', key=lambda x: x.str.lower())

        return acc_df